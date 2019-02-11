"""Usage: olop_comparison.py [options]

Compare performances of several planners on random MDPs

Options:
  -h --help
  --generate <true or false>  Generate new data [default: True].
  --show <true_or_false>      Plot results [default: True].
  --data_path <path>          Specify output data file path [default: ./out/olop_data.xlsx].
  --plot_path <path>          Specify figure data file path [default: ./out/olop_plot.png].
  --budgets <start,end,N>     Computational budgets available to planners, in logspace [default: 1,4,100].
  --samples <n>               Number of evaluations of each configuration [default: 5].
  --processes <p>             Number of processes [default: 4]
  --range <start:end>         Range of budgets to be plotted.
"""
from ast import literal_eval

from docopt import docopt
from collections import OrderedDict
from itertools import product
from multiprocessing.pool import Pool

import gym
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
import seaborn as sns

from rl_agents.agents.common import load_environment, agent_factory

gamma = 0.8
K = 5
SEED_MAX = 10000


def olop_horizon(episodes, gamma):
    return int(np.ceil(np.log(episodes) / (2 * np.log(1 / gamma))))


def allocate(budget):
    if np.size(budget) > 1:
        episodes = np.zeros(budget.shape)
        horizon = np.zeros(budget.shape)
        for i in range(budget.size):
            episodes[i], horizon[i] = allocate(budget[i])
        return episodes, horizon
    else:
        budget = np.array(budget).item()
        for episodes in range(1, budget):
            if episodes * olop_horizon(episodes, gamma) > budget:
                episodes -= 1
                break
        horizon = olop_horizon(episodes, gamma)
        return episodes, horizon


def plot_budget(budget, episodes, horizon):
    plt.figure()
    plt.subplot(311)
    plt.plot(budget, episodes)
    plt.legend(["M"])
    plt.subplot(312)
    plt.plot(budget, horizon)
    plt.legend(["L"])
    plt.subplot(313)
    plt.plot(budget, horizon / K ** (horizon - 1))
    plt.legend(['Computational complexity ratio'])
    plt.show()


def agent_configs():
    agents = {
        "random": {
            "__class__": "<class 'rl_agents.agents.simple.random.RandomUniformAgent'>"
        },
        "olop": {
            "__class__": "<class 'rl_agents.agents.tree_search.olop.OLOPAgent'>",
            "gamma": gamma,
            "max_depth": 2,
            "upper_bound": {"type": "hoeffding"},
            "lazy_tree_construction": True
        },
        "kl-olop": {
            "__class__": "<class 'rl_agents.agents.tree_search.olop.OLOPAgent'>",
            "gamma": gamma,
            "max_depth": 2,
            "upper_bound": {"type": "kullback-leibler"},
            "lazy_tree_construction": True
        }
    }
    return OrderedDict(agents)


def value_iteration():
    return {
        "__class__": "<class 'rl_agents.agents.dynamic_programming.value_iteration.ValueIterationAgent'>",
        "gamma": gamma,
        "iterations": int(3 / (1 - gamma))
    }


def evaluate(experiment):
    try:
        budget, agent_config, env_config, seed = experiment
        gym.logger.set_level(gym.logger.DISABLED)
        env = load_environment(env_config)
        env.configure({"seed": seed})
        env.seed(seed)
        state = env.reset()

        name, agent_config = agent_config
        print("Evaluating {} with budget {}".format(name, budget))
        agent_config["budget"] = int(budget)
        agent_config["iterations"] = int(env.config["max_steps"])
        agent = agent_factory(env, agent_config)
        agent.seed(seed)
        action = agent.act(state)

        values = agent_factory(env, value_iteration()).state_action_value()[env.mdp.state, :]
        result = (values[action],  np.amax(values))
        return to_dataframe(experiment, result)
    except ValueError:
        return None


def append_df_to_excel(filename=None, df=None, writer=None, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    if not writer:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        try:
            writer.book = load_workbook(filename)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            pass

    if startrow is None:
        if sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
            to_excel_kwargs["header"] = None
        else:
            startrow = 0

    if df is not None:
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
        writer.save()
        # print("Saving data to {}".format(writer.path))
    return writer


def prepare_experiments(budgets, samples):
    budgets = np.logspace(*literal_eval(budgets)).astype(int)
    envs = ['configs/FiniteMDPEnv/env_garnet.json']
    agents = agent_configs()
    seeds = np.random.randint(0, SEED_MAX, samples, dtype=int).tolist()
    experiments = list(product(budgets, agents.items(), envs, seeds))
    return experiments


def store_results(experiments, results, path=None):
    print("Finished! Exporting dataframe...")
    df = pd.DataFrame(columns=['agent', 'budget', 'value', 'optimal_value', 'seed'])
    for experiment, result in zip(experiments, results):
        _, agent, budget, seed = experiment
        value, optimal_value = result
        df = df.append({"agent": agent[0],
                        "budget": budget,
                        "seed": seed,
                        "value": value,
                        "optimal_value": optimal_value},
                       ignore_index=True)
    df["regret"] = df["optimal_value"] - df["value"]
    if path:
        append_df_to_excel(path, df, index=False)
    return df


def to_dataframe(experiment, result):
    budget, agent, _, seed = experiment
    value, optimal_value = result
    df = pd.DataFrame.from_records([{"agent": agent[0],
                                     "budget": budget,
                                     "seed": seed,
                                     "value": value,
                                     "optimal_value": optimal_value}])
    df["regret"] = df["optimal_value"] - df["value"]
    return df


def plot_all(data_path, plot_path, range):
    df = pd.read_excel(data_path)
    fig, ax = plt.subplots()
    ax.set(yscale="log")
    if range:
        start, end = range.split(':')
        df = df[df["budget"].between(int(start), int(end))]
    sns.lineplot(x="budget", y="regret", ax=ax, hue="agent", data=df, markers=True)
    print("Saving plots to {}".format(plot_path))
    plt.savefig(plot_path)


def main(args):
    if args["--generate"] == "True":
        experiments = prepare_experiments(args["--budgets"], int(args['--samples']))
        writer = append_df_to_excel(args["--data_path"])
        p = Pool(processes=int(args["--processes"]))
        print(int(args["--processes"]), "processes")
        for result in p.imap(evaluate, experiments):
            append_df_to_excel(writer=writer, df=result, index=False)
    if args["--show"] == "True":
        plot_all(args["--data_path"], args["--plot_path"], args["--range"])


if __name__ == "__main__":
    arguments = docopt(__doc__)
    main(arguments)
